
# functions for generating the intermediate 
# files based on an Excel document 

import openpyxl 
import os 
import sys 
import csv 


class ExcelScanResults:

    def __init__(self):
        self.formulas = dict()
        self.formulas["$$$GLOBAL$$$"] = dict()
        self.assigned_indexes = dict()
        self.const = dict()
        self.const["$$$GLOBAL$$$"] = dict()
        self.const_indexes = dict()



    def record_formula(self, formula_str, in_cell_or_name, in_sheet):
        
        if in_sheet is not None:
        
            if in_sheet not in self.formulas:
                self.formulas[in_sheet] = dict()
            
            self.formulas[in_sheet][in_cell_or_name] = formula_str
        
        else:
            # if not in a sheet, it's global 
            self.formulas["$$$GLOBAL$$$"][in_cell_or_name] = formula_str


    def record_const(self, s, in_cell_or_name, in_sheet):
        
        if in_sheet is not None:
        
            if in_sheet not in self.const:
                self.const[in_sheet] = dict()
            
            self.const[in_sheet][in_cell_or_name] = s
        
        else:
            # if not in a sheet, it's global 
            self.const["$$$GLOBAL$$$"][in_cell_or_name] = s
            

    def assign_formula_indexes(self):
        _id = 1
        for scope in self.formulas:
            for nm in self.formulas[scope]:
                t = (scope, nm)
                self.assigned_indexes[t] = _id 
                _id = _id + 1 
        
        _id = 1
        for scope in self.const:
            for nm in self.const[scope]:
                t = (scope, nm)
                self.const_indexes[t] = _id 
                _id = _id + 1 
        




def is_excel_formula(cell_value):
    return str(cell_value).startswith("=")
        


def scan_excel(input_excel):
    
    scan_r = ExcelScanResults()
    
    # thanks https://stackoverflow.com/questions/13377793/is-it-possible-to-get-an-excel-documents-row-count-without-loading-the-entire-d
    wb = openpyxl.load_workbook(input_excel)
    print("scanning " + input_excel)
    sheets = wb.sheetnames 

    sheet_idx = 1

    sheet_idx_to_sheet_name = dict()

    for sheet_name in sheets:
        
        # print("" + str(sheet_idx) + " " + sheet_name)
        sheet_idx_to_sheet_name[sheet_idx] = sheet_name
        sht = wb[sheet_name]
        
        for row in sht.iter_rows(min_row=1, max_col=sht.max_column, max_row=sht.max_row):
            for cell in row:
                if cell.value is not None:
                    if is_excel_formula(cell.value):
                        scan_r.record_formula(cell.value, cell.coordinate, sheet_name)
                    else:
                        if cell.value != "":
                            scan_r.record_const(cell.value, cell.coordinate, sheet_name)
        
        sheet_idx = sheet_idx + 1


    # print("sheet index to sheet name")
    # for idx in sheet_idx_to_sheet_name:
    #     print(str(idx) + " = " + str(sheet_idx_to_sheet_name[idx]))

    # print("defined names at the workbook level ")
    for dn in wb.defined_names.definedName:
        if dn.localSheetId is not None:
            sheet_scope = sheet_idx_to_sheet_name[dn.localSheetId]
        else:
            sheet_scope = None 

        scan_r.record_formula(dn.attr_text, dn.name, sheet_scope)

    print("Done scanning ")

    scan_r.assign_formula_indexes()  # ensures formulas are unique 

    return scan_r
            


def dump_scanned_formulas(scan_r, folder):   # input ScanResults to this one 
    formulas_file = dump_scanned_formulas_path(folder)
    with open(formulas_file, "w", newline="") as fp:
        wr = csv.writer(fp)
        wr.writerow(["formula_id", "sheet", "cell_or_name", "formula"])
        for scope in scan_r.formulas:
            for nm in scan_r.formulas[scope]:
                t = (scope, nm)
                _id = scan_r.assigned_indexes[t]
                wr.writerow([_id, scope, nm, scan_r.formulas[scope][nm]])



def dump_scanned_constants(scan_r, folder):   # input ScanResults to this one 
    const_file = dump_scanned_constants_path(folder)
    with open(const_file, "w", newline="") as fp:
        wr = csv.writer(fp)
        wr.writerow(["const_id", "sheet", "cell_or_name", "value"])
        for scope in scan_r.const:
            for nm in scan_r.const[scope]:
                t = (scope, nm)
                _id = scan_r.const_indexes[t]
                wr.writerow([_id, scope, nm, scan_r.const[scope][nm]])


    
def dump_scanned_formulas_path(folder):
    return os.path.join(folder, "excel_formulas.csv")




def dump_scanned_constants_path(folder):
    return os.path.join(folder, "excel_constants.csv")


